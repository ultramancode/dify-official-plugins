"""Abstract interface for document loader implementations."""

import os
from typing import Optional, cast

import pandas as pd
from openpyxl import load_workbook  # type: ignore

from tools.extractor_base import BaseExtractor
from tools.document import Document, ExtractorResult
from io import BytesIO


class ExcelExtractor(BaseExtractor):
    """Load Excel files.


    Args:
        file_bytes: file bytes.
        file_name: file name.
        encoding: encoding.
        autodetect_encoding: autodetect encoding.
    """

    def __init__(
        self,
        file_bytes: bytes,
        file_name: str,
        encoding: Optional[str] = None,
        autodetect_encoding: bool = False,
    ):
        """Initialize with file path."""
        self._file_bytes = file_bytes
        self._file_name = file_name
        self._encoding = encoding
        self._autodetect_encoding = autodetect_encoding

    def extract(self) -> ExtractorResult:
        """Load from Excel file in xls or xlsx format using Pandas and openpyxl."""
        documents = []
        all_md_content = ""
        file_extension = os.path.splitext(self._file_name)[-1].lower()

        if file_extension == ".xlsx":
            wb = load_workbook(filename=BytesIO(self._file_bytes), data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                data = sheet.values
                try:
                    cols = next(data)
                except StopIteration:
                    continue
                df = pd.DataFrame(data, columns=cols)

                df.dropna(how="all", inplace=True)

                md_content = f"## {sheet_name}\n\n"
                md_content += "| " + " | ".join(str(col) for col in cols) + " |\n"
                md_content += "| " + " | ".join(["---"] * len(cols)) + " |\n"

                for index, row in df.iterrows():
                    md_content += (
                        "| "
                        + " | ".join(
                            [str(cell) if pd.notna(cell) else "" for cell in row]
                        )
                        + " |\n"
                    )

                    page_content = []
                    for col_index, (k, v) in enumerate(row.items()):
                        if pd.notna(v):
                            cell = sheet.cell(
                                row=cast(int, index) + 2, column=col_index + 1
                            )  # +2 to account for header and 1-based index
                            if cell.hyperlink:
                                value = f"[{v}]({cell.hyperlink.target})"
                                page_content.append(f'"{k}":"{value}"')
                            else:
                                page_content.append(f'"{k}":"{v}"')
                    documents.append(
                        Document(
                            page_content=";".join(page_content),
                            metadata={"source": self._file_name},
                        )
                    )
                all_md_content += md_content + "\n"

        elif file_extension == ".xls":
            excel_file = pd.ExcelFile(self._file_bytes, engine="xlrd")
            for excel_sheet_name in excel_file.sheet_names:
                df = excel_file.parse(sheet_name=excel_sheet_name)
                df.dropna(how="all", inplace=True)

                md_content = f"## {excel_sheet_name}\n\n"
                md_content += "| " + " | ".join(df.columns) + " |\n"
                md_content += "| " + " | ".join(["---"] * len(df.columns)) + " |\n"

                for _, row in df.iterrows():
                    md_content += (
                        "| "
                        + " | ".join(
                            [str(cell) if pd.notna(cell) else "" for cell in row]
                        )
                        + " |\n"
                    )
                    page_content = []
                    for k, v in row.items():
                        if pd.notna(v):
                            page_content.append(f'"{k}":"{v}"')
                    documents.append(
                        Document(
                            page_content=";".join(page_content),
                            metadata={"source": self._file_name},
                        )
                    )
                all_md_content += md_content + "\n"
        else:
            raise ValueError(f"Unsupported file extension: {file_extension}")

        return ExtractorResult(md_content=all_md_content, documents=documents)
